import logging

from RPA.Browser.Selenium import Selenium
from RPA.HTTP import HTTP
from openpyxl.workbook import Workbook
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from RPA.Archive import Archive
import shutil

from constants import JS_SCRIPT, SORT_BY, SITE_URL, STATIC, FILES_DIRECTORY, ZIP_FILE
from utils import XPATH_CONFIG, is_amount_mentioned


class CollectFreshNews:
    def __init__(self, query):
        self.browser = Selenium()
        self.search_query = query
        self.http_lib = HTTP()
        self.archive = Archive()
        self.news_data = []

    def open_available_browser(self, url):
        """Opens an available browser and navigates to the given URL"""
        try:
            self.browser.open_available_browser(url)
            logging.info(f"Opened browser and navigated to URL: {url}")
        except Exception as e:
            logging.error(f"Error opening browser: {e}")

    def close_browser(self):
        self.browser.close_browser()

    def click_search_and_enter_query(self):
        """Clicks the search button and enters the search query"""
        try:
            logging.info(f"Starting entering search query: {self.search_query}")
            self.browser.wait_until_page_contains_element(XPATH_CONFIG.get('search_button'))
            self.browser.click_button(XPATH_CONFIG.get('search_button'))
            self.browser.input_text(XPATH_CONFIG.get('search_input'), self.search_query)
            self.browser.click_button(XPATH_CONFIG.get('submit_search'))
            self.browser.wait_until_element_is_visible(XPATH_CONFIG.get('search_result_list'), timeout=20)
            logging.info("Clicked search button and entered search query.")
        except Exception as e:
            logging.error(f"Error clicking search button: {e}")

    def check_and_get_more_data(self):
        """Checks for more data and fetches it if available"""
        try:
            show_more = self.browser.is_element_visible(XPATH_CONFIG.get('show_more'))
            if show_more:
                self.browser.wait_until_page_contains_element(XPATH_CONFIG.get('show_more'))
                self.browser.execute_javascript(JS_SCRIPT)
                self.browser.click_element_when_clickable(XPATH_CONFIG.get('show_more'))
                self.browser.wait_until_element_is_not_visible(XPATH_CONFIG.get('loading_show_more'), timeout=10)
                more_data = self.browser.is_element_visible(XPATH_CONFIG.get('show_more'))
                if more_data:
                    self.check_and_get_more_data()
        except Exception as e:
            logging.error(f"Error getting more data: {e}")

    def apply_sort_by_value(self, sorter):
        """Applies the specified sorting method"""
        try:
            logging.error(f"Applying sorter for {sorter}")
            self.browser.select_from_list_by_value(XPATH_CONFIG.get('sorter'), sorter)
            self.browser.wait_until_page_contains_element(XPATH_CONFIG.get('search_result_list'))
        except Exception as e:
            logging.error(f"Error applying sort by value: {e}")

    def count_search_phrases(self, title, description):
        """Counts the number of search phrases in the title and description"""
        try:
            title_search_phrase_count = title.lower().count(self.search_query.lower())
            description_search_phrase_count = description.lower().count(self.search_query.lower())
            return int(title_search_phrase_count) + int(description_search_phrase_count)
        except Exception as e:
            logging.error(f"Error counting search phrases: {e}")
            return 0

    def download_image(self, url, title):
        """Downloads an image from the specified URL"""
        try:
            image_name = f"{STATIC}{FILES_DIRECTORY}/{title}.jpg"
            self.http_lib.download(url=url, target_file=image_name)
            return image_name
        except FileNotFoundError:
            logging.error(f"Error downloading image for : {title}")
            return None

    def extract_data_from_list(self):
        """Extracts data from the list of articles"""
        try:
            articles = self.browser.find_elements(XPATH_CONFIG.get('article'))
            logging.info(f"Total {len(articles)} are available to scrap data.")
            for index, article in enumerate(articles):
                try:
                    title = article.find_element(by=By.XPATH, value=XPATH_CONFIG.get('title')).text
                    title = title.replace("...", "").strip()
                    logging.info(f"Extracting data for: {title}.")
                    description = article.find_element(by=By.XPATH, value=XPATH_CONFIG.get('description')).text
                    filename = article.find_element(by=By.XPATH, value=XPATH_CONFIG.get('filename')).get_attribute(
                        'src')
                    count_phrases = self.count_search_phrases(title, description)
                    includes_amount = is_amount_mentioned(title + description)
                    image = self.download_image(filename, title)

                    try:
                        date = article.find_element(by=By.XPATH, value=XPATH_CONFIG.get('date')).text
                    except NoSuchElementException:
                        date = ''

                    self.news_data.append({
                        'Title': title,
                        'Date': date,
                        'Description': description,
                        'Image': image,
                        'Count phrase': count_phrases,
                        'Includes Amount': includes_amount
                    })
                except NoSuchElementException:
                    logging.error(f"Error while extracting {index}th article")
        except Exception as e:
            logging.error(f"Error extracting data from list: {e}")

    def save_data_in_excel_file(self, filename):
        """Saves the extracted data in an Excel file"""
        try:
            logging.info("Starting exporting data to excel file.")
            workbook = Workbook()
            sheet = workbook.active
            header = ['Title', 'Date', 'Description', 'Image', 'Count phrase', 'Includes Amount']
            for col_num, key in enumerate(header, start=1):
                sheet.cell(row=1, column=col_num, value=key)
            for i, row in enumerate(self.news_data, start=2):
                for j, key in enumerate(header):
                    sheet.cell(row=i, column=j + 1, value=row[key])
            workbook.save(f"{STATIC}{filename}.xlsx")
            logging.info("Excel file downloaded successfully.")
        except Exception as e:
            logging.error(f"Error saving data in excel: {e}")

    def archive_folder_and_remove_images_directory(self):
        """Moving images to zip folder to minimize the file size in control room"""
        try:
            logging.info("Starting archive and removing images directory.")
            folder_path = f'{STATIC}{FILES_DIRECTORY}'
            self.archive.archive_folder_with_zip(
                include='*.jpg',
                folder=folder_path,
                archive_name=f'{STATIC}{ZIP_FILE}.zip'
            )
            shutil.rmtree(folder_path)
        except FileNotFoundError as e:
            logging.error(f"Error: Directory not found: {e}")
        except Exception as e:
            logging.error(f"Error while archiving and removing images directory: {e}")

    def task(self):
        """Insert the sales data for the week and export it as a PDF"""
        try:
            logging.info("Starting scrapper for fresh news.")
            self.open_available_browser(SITE_URL)
            self.click_search_and_enter_query()

            is_no_data_indication = self.browser.is_element_visible(XPATH_CONFIG.get('no_result_indication'))
            if is_no_data_indication:
                logging.info("No data found for given query.")
            else:
                self.apply_sort_by_value(SORT_BY)
                self.check_and_get_more_data()
                self.extract_data_from_list()
                self.save_data_in_excel_file('exported_data')
                self.archive_folder_and_remove_images_directory()
            self.close_browser()
        except Exception as e:
            logging.error(f"Error executing task: {e}")
